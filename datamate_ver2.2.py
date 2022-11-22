import sys
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.QtGui import *
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from string import ascii_lowercase
import os
from scipy import stats
import pandas as pd
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows


form_class = uic.loadUiType("datamate_ver2.2.ui")[0]
print("""
Datamate_ver2.2 made by 이성우
잠시만 기다려주세요...
""")


class WindowClass(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowIcon(QIcon('icon.png'))

        self.lineEdit.setValidator(QIntValidator(1, 1000))
        self.lineEdit_2.setValidator(QIntValidator(1, 1000))

        self.tableWidget.setItem(0, 0, QTableWidgetItem("더블 클릭"))
        self.tableWidget.setItem(0, 1, QTableWidgetItem("숫자만 입력"))
        self.tableWidget.setItem(0, 2, QTableWidgetItem("숫자만 입력"))

        self.lineEdit_3.setEnabled(False)
        self.btn_rtn.clicked.connect(self.btn_rtn_function)  # 입력값을 받아서 데이터시트 생성 메서드로 전달
        self.btn_fload.clicked.connect(self.btn_fload_function)
        self.btn_anal.setEnabled(False)
        self.btn_anal.clicked.connect(self.btn_anal_function)
        self.rbtn_halftest.clicked.connect(self.test_indicator)
        self.rbtn_grouptest.clicked.connect(self.test_indicator)
        self.rbtn_halftest.setEnabled(0)
        self.rbtn_grouptest.setEnabled(0)
        self.csv = 0
        self.chbx_csv.stateChanged.connect(self.csv_indicator)

        self.listview = QListView()
        self.model = QStandardItemModel()


    def csv_indicator(self):
        if self.chbx_csv.isChecked():
            self.csv = 1
        else:
            self.csv = 0

    def test_indicator(self):
        self.btn_anal.setEnabled(True)
        if self.rbtn_halftest.isChecked() : self.det = 0
        elif self.rbtn_grouptest.isChecked() : self.det = 1


    def btn_rtn_function(self):

        if self.lineEdit.text() and self.lineEdit_2.text() and self.tableWidget.item(0, 0) \
                and self.tableWidget.item(0, 1) and self.tableWidget.item(0, 2) \
                and [self.tableWidget.item(i, 1).text().isdecimal() for i in range(self.tableWidget.rowCount())
                     if self.tableWidget.item(i, 1)].count(False) == 0 \
                and [self.tableWidget.item(i, 2).text().isdecimal() for i in range(self.tableWidget.rowCount())
                     if self.tableWidget.item(i, 2)].count(False) == 0:

            sub = int(self.lineEdit.text())  # 입력 받은 실험 조건들을 변수, 리스트 요소로 선언
            t = int(self.lineEdit_2.text())
            list_para = [self.tableWidget.item(i, 0).text() for i in range(self.tableWidget.rowCount()) if
                         self.tableWidget.item(i, 0)]
            list_sites = [int(self.tableWidget.item(i, 1).text()) for i in range(self.tableWidget.rowCount()) if
                          self.tableWidget.item(i, 1)]
            list_trials = [int(self.tableWidget.item(i, 2).text()) for i in range(self.tableWidget.rowCount()) if
                           self.tableWidget.item(i, 2)]

            if len(list_para) == len(list_sites) == len(list_trials):  # 제대로 입력이 되었을 때만 데이터 시트 생성

                wb = Workbook()
                ws1 = wb.active
                ws1.title = 'input'

                # 기본 헤더 정의
                ws1['A1'].value = 'dm2.2'
                ws1['A2'].value = '이름'
                ws1['B2'].value = '나이'
                ws1['C2'].value = 'M/F'
                ws1['D2'].value = 'parameter'
                ws1['E2'].value = '부위'
                ws1['F2'].value = '사용 제품'
                ws1.freeze_panes = 'G3'

                # defining rows
                parameters = len(list_para)
                thick = Side(border_style='thick', color='000000')
                thin = Side(border_style='thin', color='000000')

                for i in range(sub):
                    ws1.merge_cells(start_row=3 + sum(list_sites) * i, start_column=1,
                                    end_row=2 + sum(list_sites) * (i + 1), end_column=1)
                    ws1.merge_cells(start_row=3 + sum(list_sites) * i, start_column=2,
                                    end_row=2 + sum(list_sites) * (i + 1), end_column=2)
                    ws1.merge_cells(start_row=3 + sum(list_sites) * i, start_column=3,
                                    end_row=2 + sum(list_sites) * (i + 1), end_column=3)

                    ws1.cell(column=1, row=3 + sum(list_sites) * i).alignment = Alignment(horizontal='center',
                                                                                          vertical='center')
                    ws1.cell(column=2, row=3 + sum(list_sites) * i).alignment = Alignment(horizontal='center',
                                                                                          vertical='center')
                    ws1.cell(column=3, row=3 + sum(list_sites) * i).alignment = Alignment(horizontal='center',
                                                                                          vertical='center')

                    for j in range(6 + (max(list_trials) + 1) * t):
                        ws1.cell(column=j + 1, row=3 + sum(list_sites) * i).border = Border(top=thick)
                        if i == sub - 1:
                            ws1.cell(column=j + 1, row=3 + sum(list_sites) * sub).border = Border(top=thick)

                    n = 0  # parameter간의 간격
                    for parameter in range(parameters):
                        ws1.cell(column=4, row=3 + sum(list_sites) * i + n).value = list_para[parameter]
                        ws1.merge_cells(start_row=3 + sum(list_sites) * i + n, start_column=4,
                                        end_row=2 + sum(list_sites) * i + n + list_sites[parameter], end_column=4)
                        ws1.cell(column=4, row=3 + sum(list_sites) * i + n).alignment = Alignment(horizontal='center',
                                                                                                  vertical='center')
                        if parameter != 0:
                            for j in range(4, 7 + (max(list_trials) + 1) * t):
                                ws1.cell(column=j, row=3 + sum(list_sites) * i + n).border = Border(top=thin)

                        for k in range(list_sites[parameter]):
                            ws1.cell(column=5, row=3 + sum(list_sites) * i + n + k).value = list(ascii_lowercase)[k]
                            if max(list_trials) - list_trials[parameter]:  # 쓰이지 않을 셀에 색 채우기
                                for j in range(t):
                                    for trial in range(max(list_trials) - list_trials[parameter]):
                                        ws1.cell(
                                            column=6 + list_trials[parameter] + trial + 1 + (max(list_trials) + 1) * j,
                                            row=3 + sum(list_sites) * i + n + k).fill = PatternFill('solid',
                                                                                                    fgColor='D9D9D9')
                        n += list_sites[parameter]

                # defining columns
                import math
                ordinal = lambda n: "%d%s" % (n, "tsnrhtdd"[(math.floor(n / 10) % 10 != 1) * (n % 10 < 4) * n % 10::4])
                list_col = list(ascii_lowercase) + [''.join(i) for i in
                                                    [(a, b) for a in list(ascii_lowercase) for b in
                                                     list(ascii_lowercase)]]

                for i in range(t):
                    if i == 0:
                        ws1.cell(row=1, column=7).value = '전'
                    ws1.merge_cells(start_column=7 + (max(list_trials) + 1) * i, start_row=1,
                                    end_column=6 + (max(list_trials) + 1) * (i + 1), end_row=1)
                    ws1.cell(row=1, column=7 + (max(list_trials) + 1) * i).alignment = Alignment(horizontal='center',
                                                                                                 vertical='center')
                    for k in range(max(list_trials) + 1):
                        ws1.cell(row=2, column=7 + (max(list_trials) + 1) * i + k).value = ordinal(k + 1)
                        if k == max(list_trials):
                            col_avg = 7 + (max(list_trials) + 1) * i + k - 1
                            ws1.cell(row=2, column=col_avg + 1).value = 't_%d_avg' % i
                            ws1.cell(row=2, column=col_avg + 1).font = Font(bold=True)
                            for j in range(sub * sum(list_sites)):
                                ws1.cell(row=3 + j, column=col_avg + 1).value = '=average(%s%d:%s%d)' % (
                                    list_col[col_avg - max(list_trials)], 3 + j, list_col[col_avg - 1], 3 + j)
                                ws1.cell(row=3 + j, column=col_avg + 1).font = Font(bold=True)

                wb.save('data_input.xlsx')

                # 사용자 편의를 위해 바로 데이터 시트 띄워주기
                os.startfile("data_input.xlsx")

                self.lineEdit.clear()  # 입력 필드 초기화
                self.lineEdit_2.clear()
                self.tableWidget.clearContents()
                self.tableWidget.setItem(0, 0, QTableWidgetItem("더블 클릭"))
                self.tableWidget.setItem(0, 1, QTableWidgetItem("숫자만 입력"))
                self.tableWidget.setItem(0, 2, QTableWidgetItem("숫자만 입력"))

            else:
                QMessageBox.about(self, "안내", "parameter 필드를 정확히 입력해주세요")

            # parameter 입력 필드 초기화
            self.tableWidget.clearContents()
            self.tableWidget.setItem(0, 0, QTableWidgetItem("더블 클릭"))
            self.tableWidget.setItem(0, 1, QTableWidgetItem("숫자만 입력"))
            self.tableWidget.setItem(0, 2, QTableWidgetItem("숫자만 입력"))

        else:
            QMessageBox.about(self, "안내", "모든 필드에 값을 정확히 입력해주세요")

    def btn_fload_function(self):
        fname = QFileDialog.getOpenFileName(self, 'Open file',  os.path.join(os.path.expanduser('~'),'Desktop'),"excel file(*.xlsx *.xlam)")

        if fname[0]:
            self.fname = fname
            df_load = pd.read_excel(self.fname[0],  header = None, nrows = 2)
            df_load = df_load.where(pd.notnull(df_load), None)
            if df_load.iat[0,0] != 'dm2.2':
                QMessageBox.about(self, "경고", "적합한 파일을 선택해주세요")
                self.lineEdit_3.clear()
                self.rbtn_halftest.setEnabled(0)
                self.rbtn_grouptest.setEnabled(0)
                return

            self.trial = [i for i in df_load.iloc[1, :]].index('t_0_avg') - 6
            self.dict_time = {}
            for i in range(1, df_load.columns.size):
                if df_load.iloc[0, i] != None:
                    self.dict_time[df_load.iloc[1, i + self.trial]] = str(df_load.iloc[0, i])
            print(self.trial)
            print(self.dict_time, len(self.dict_time))

            if len([i for i in df_load.iloc[0] if not i == None]) == 0:
                QMessageBox.about(self, "경고", "측정 시점을 작성해주세요")
                self.lineEdit_3.clear()
                self.rbtn_halftest.setEnabled(0)
                self.rbtn_grouptest.setEnabled(0)
                self.btn_anal.setEnabled(False)

            elif len([a for a in df_load.iloc[1] if 'avg' in a]) != len(self.dict_time):
                text = QMessageBox.question(self, "안내", """%s 모든 측정 시점이 맞습니까?""" %[str(i) for i in self.dict_time.values()], \
                                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

                if text == QMessageBox.Yes:
                    self.lineEdit_3.setText(fname[0])
                    self.rbtn_halftest.setEnabled(1)
                    self.rbtn_grouptest.setEnabled(1)
                    self.parameters = np.unique(pd.read_excel(self.fname[0], header=1)['parameter'].dropna(axis=0))
                    for para in self.parameters:
                        item = QStandardItem(para)
                        item.setCheckable(True)
                        self.model.appendRow(item)

                    self.listView.setModel(self.model)

                else:
                    self.lineEdit_3.clear()
                    self.rbtn_halftest.setEnabled(0)
                    self.rbtn_grouptest.setEnabled(0)
                    self.btn_anal.setEnabled(False)
            else:
                self.lineEdit_3.setText(fname[0])
                self.rbtn_halftest.setEnabled(1)
                self.rbtn_grouptest.setEnabled(1)
                self.parameters = np.unique(pd.read_excel(self.fname[0], header=1)['parameter'].dropna(axis=0))
                for para in self.parameters:
                    item = QStandardItem(para)
                    item.setCheckable(True)
                    self.model.appendRow(item)

                self.listView.setModel(self.model)


        else:
            pass

    def btn_anal_function(self):
        rc =[]
        for i in range(len(self.parameters)):
            if self.model.item(i).checkState():
                rc.append(self.model.item(i).text())

        df = pd.read_excel(self.fname[0], header=1)
        df = df.where(pd.notnull(df), None)

        sub = np.array([i for i in df.iloc[:, 0] if not i == None])
        line = len([i for i in df.iloc[:, 4] if i])
        ages = np.array([df.iat[i, 1] for i in range(line) if df.iat[i, 0] != None])
        sex = np.array([df.iat[i, 2] for i in range(line) if df.iat[i, 0] != None])
        width = df.columns.size

        n = 0
        for i in range(0, line):
            if not df.iat[i, 0] == None:
                df.iat[i, 0] = n
                n += 1

        for i in range(line):
            if df.iloc[i, 0] == None:
                df.iloc[i, 0] = df.iloc[i - 1, 0]
                df.iloc[i, 1] = df.iloc[i - 1, 1]
                df.iloc[i, 2] = df.iloc[i - 1, 2]
            if df.iloc[i, 3] == None:
                df.iloc[i, 3] = df.iloc[i - 1, 3]

        df2 = df.melt(id_vars=df.columns[:6], value_vars=self.dict_time.keys(), var_name='time')
        df2['value'] = pd.to_numeric(df2['value'])

        wb = Workbook()
        ws = wb.active
        ws.title = 'info'
        ws['A1'].value = '#'
        ws['B1'].value = '이름'
        ws['C1'].value = '나이'
        ws['D1'].value = '성별'


        for i in range(len(sub)):
            ws[2 + i][0].value = i + 1
            ws[2 + i][1].value = sub[i]
            ws[2 + i][2].value = ages[i]
            ws[2 + i][3].value = sex[i]



        if self.det == 0:
            order = 0
            for para in self.parameters:
                para = wb.create_sheet(para)

                piv_table0 = pd.pivot_table(df2.loc[df2.parameter == self.parameters[order]], index=df2.columns[0],
                                            columns=[df2.columns[5], df2.columns[6]], values='value')
                anal_pdct = np.sort(np.unique([a for a, b in piv_table0.columns]))
                anal_time = np.array([self.dict_time[i] for i in np.sort(np.unique([b for a, b in piv_table0.columns]))])

                piv_table0.columns = np.array(['%s_%s' % (a, self.dict_time[b]) for a, b in piv_table0.columns])
                piv_table0.index += 1
                if self.parameters[order] in rc:
                    rcv_table = pd.DataFrame(index=piv_table0.index)
                    for pdct in anal_pdct:
                        for time in anal_time[2:]:
                            temp_series = (piv_table0.loc[:, ('%s_%s' % (pdct, anal_time[1]))] - piv_table0.loc[:, ('%s_%s' %(pdct, time))]) / \
                                          (piv_table0.loc[:, ('%s_%s' % (pdct, anal_time[1]))] - piv_table0.loc[:, ('%s_%s' %(pdct, anal_time[0]))]) * 100
                            temp_series = temp_series.rename('%s_%s(rc)' % (pdct, time))
                            rcv_table = pd.concat([rcv_table, temp_series], axis=1)
                    piv_table0 = rcv_table

                mn = piv_table0.mean()
                piv_table = piv_table0.append([mn, mn, mn], ignore_index=True)

                sw = pd.Series([stats.shapiro(np.array([v for v in piv_table0[i] if np.isnan(v) == False]))[1] \
                                    if len(np.array([v for v in piv_table0[i] if np.isnan(v) == False])) > 3 else None
                                for i in piv_table0])
                ks = pd.Series([stats.kstest(np.array([v for v in piv_table0[i] if np.isnan(v) == False]), 'norm', \
                                             args=(
                                             np.mean(np.array([v for v in piv_table0[i] if np.isnan(v) == False])), \
                                             np.std(np.array([v for v in piv_table0[i] if np.isnan(v) == False]))))[1] \
                                    if len(np.array([v for v in piv_table0[i] if np.isnan(v) == False])) > 3 else None
                                for i in piv_table0])

                piv_table.index = list(piv_table0.index) + ['mean', 's-w', 'k-s']

                for i in range(len(piv_table.loc['mean'])):
                    piv_table.loc['s-w'][i] = sw[i]
                    piv_table.loc['k-s'][i] = ks[i]

                for r in dataframe_to_rows(piv_table, index=True, header=True):
                    para.append(r)

                for i in range(anal_pdct.size):
                    para.cell(row=8 + piv_table0.index.size + i, column=1).value = anal_pdct[i]
                    if self.parameters[order] in rc:
                        for j in range(anal_time.size-2):
                            para.cell(row=8 + piv_table0.index.size + i, column = 2 + j).value = \
                                piv_table.loc['mean', '%s_%s(rc)' %(anal_pdct[i], anal_time[j+2])]
                            if i == 0:
                                para.cell(row = 7 + piv_table0.index.size, column=2 + j).value = anal_time[j+2]+'(rc)'
                    else:
                        for j in range(anal_time.size):
                            para.cell(row=8 + piv_table0.index.size + i, column=2 + j).value = \
                                piv_table.loc['mean', '%s_%s' % (anal_pdct[i], anal_time[j])]
                            if i == 0:
                                para.cell(row=7 + piv_table0.index.size, column=2 + j).value = anal_time[j]

                for cell in para[piv_table0.index.size + 3][0:]:
                    cell.font = Font(bold=True)
                for cell in [i if i.value else None for i in para[piv_table0.index.size + 4][1:]]:
                    if cell.value < 0.05:
                        cell.fill = PatternFill('solid', fgColor='00FFFF00')
                for cell in [i if i.value else None for i in para[piv_table0.index.size + 5][1:]]:
                    if cell.value < 0.05:
                        cell.fill = PatternFill('solid', fgColor='00FF0000')

                if self.csv:
                    piv_table0.to_csv(os.path.dirname(self.fname[0])+'/%s.csv' % self.parameters[order], encoding = 'CP949')

                order += 1

        elif self.det == 1:
            order = 0
            for para in self.parameters:
                para = wb.create_sheet(para)
                df_group = df2.loc[df2.parameter == self.parameters[order]]
                gr0 = pd.pivot_table(df_group, index=df2.columns[0], \
                                     columns=[df2.columns[5], df2.columns[6]], values='value')
                groups = np.unique(df_group[df2.columns[5]].dropna())

                order_group = 0
                for group in groups:
                    pvt = gr0[(group,)].dropna(thresh=gr0[(group,)].shape[1])
                    anal_time = pvt.columns

                    if self.parameters[order] in rc:
                        rcv_table = pd.DataFrame(index = pvt.index)
                        for time in anal_time[2:]:
                            temp_series = (pvt.loc[:, '%s' %anal_time[1]] - pvt.loc[:, '%s' %time])/ \
                                          (pvt.loc[:, '%s' %anal_time[1]] - pvt.loc[:, '%s' %anal_time[0]])*100
                            temp_series = temp_series.rename('%s' %time)
                            rcv_table = pd.concat([rcv_table, temp_series], axis = 1)
                        pvt = rcv_table

                    mn2 = pvt.mean()
                    pvt1 = pvt.append([mn2, mn2, mn2], ignore_index=True)

                    sw = pd.Series([stats.shapiro(np.array([v for v in pvt[i] if np.isnan(v) == False]))[1] \
                                        if len(np.array([v for v in pvt[i] if np.isnan(v) == False])) > 3 else None for
                                    i in pvt])
                    ks = pd.Series([stats.kstest(np.array([v for v in pvt[i] if np.isnan(v) == False]), 'norm', \
                                                 args=(np.mean(np.array([v for v in pvt[i] if np.isnan(v) == False])), \
                                                       np.std(np.array([v for v in pvt[i] if np.isnan(v) == False]))))[
                                        1] \
                                        if len(np.array([v for v in pvt[i] if np.isnan(v) == False])) > 3 else None for
                                    i in pvt])

                    pvt1.index = list(np.array(gr0[(group,)].dropna().index) + 1) + ['mean', 's-w', 'k-s']
                    for i in range(len(pvt1.loc['mean'])):
                        pvt1.loc['s-w'][i] = sw[i]
                        pvt1.loc['k-s'][i] = ks[i]

                    if self.parameters[order] in rc:
                        pvt1.columns = np.array([self.dict_time[com]+'(rc)' for com in pvt1.columns])
                    else:
                        pvt1.columns = np.array([self.dict_time[com] for com in pvt1.columns])

                    para.append([group])

                    for r in dataframe_to_rows(pvt1, index=True, header=True):
                        para.append(r)

                    para.append([])
                    for i in range(pvt1.columns.size):
                        para.cell(row=4 + order_group, column=4 + pvt1.columns.size + i).value = np.array(mn2)[i]

                    order_group += 1

                for i in range(len(groups)):
                    para.cell(row=4 + i, column=3 + pvt1.columns.size).value = groups[i]
                    if i == 0:
                        for j in range(pvt1.columns.size):
                            para.cell(row=3, column=4 + pvt1.columns.size + j).value = pvt1.columns[j]

                for r in np.array([cell.row for cell in para['A'] if cell.value == 'mean']):
                    for c in range(1, pvt1.columns.size + 2):
                        para.cell(row=r, column=c).font = Font(bold=True)
                        if c != 1:
                            if para.cell(row=r + 1, column=c).value < 0.05:
                                para.cell(row=r + 1, column=c).fill = PatternFill('solid', fgColor='00FFFF00')
                            if para.cell(row=r + 2, column=c).value < 0.05:
                                para.cell(row=r + 2, column=c).fill = PatternFill('solid', fgColor='00FF0000')

                if self.csv:
                    df_group.loc[:, df2.columns[0]] += 1
                    df_csv = pd.pivot_table(df_group, index=[df2.columns[5], df.columns[0]], columns='time', values='value')
                    df_csv.columns = np.array([self.dict_time[a] for a in df_csv.columns])
                    if self.parameters[order] in rc:
                        rcv_table = pd.DataFrame(index=df_csv.index)
                        for time in df_csv.columns[2:]:
                            temp_series = (df_csv.loc[:, '%s' % df_csv.columns[1]] - df_csv.loc[:, '%s' % time]) / \
                                          (df_csv.loc[:, '%s' % df_csv.columns[1]] - df_csv.loc[:, '%s' % df_csv.columns[0]]) * 100
                            temp_series = temp_series.rename('%s(rc)' % time)
                            rcv_table = pd.concat([rcv_table, temp_series], axis=1)
                        df_csv = rcv_table
                    df_csv.to_csv(os.path.dirname(self.fname[0])+'/%s.csv' % self.parameters[order], encoding = 'CP949')

                order += 1

        ws_input = wb.create_sheet('input')
        df.iloc[:, 0] = df.iloc[:, 0] + 1
        df.columns = np.array([self.dict_time[column] if column in self.dict_time.keys() and 'avg' in column else column for column in df.columns])
        df_out = df[[i for i in df.columns[:6]]+[str(i) for i in self.dict_time.values()]]
        for r in dataframe_to_rows(df_out, index=False, header=True):
            ws_input.append(r)

        wb.save(os.path.dirname(self.fname[0])+'/data_result.xlsx')
        os.startfile(os.path.dirname(self.fname[0])+'/data_result.xlsx')


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    app.exec_()
